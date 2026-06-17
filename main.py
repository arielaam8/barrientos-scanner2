import os
import threading
from kivy.app import App
from kivy.uix.screenmanager import ScreenManager, Screen, SlideTransition
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from kivy.uix.popup import Popup
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.camera import Camera
from kivy.clock import Clock
from kivy.utils import platform
from kivy.graphics import Color, RoundedRectangle
from kivy.core.window import Window
from kivy.metrics import dp

import openpyxl

if platform == 'android':
    from android.permissions import request_permissions, Permission
    from android.storage import primary_external_storage_path

# ── Colors ──────────────────────────────────────────────────────────────────
BG      = (0.10, 0.10, 0.18, 1)
PANEL   = (0.09, 0.13, 0.24, 1)
ACCENT  = (0.91, 0.27, 0.38, 1)
GREEN   = (0.11, 0.31, 0.24, 1)
TEXT    = (0.93, 0.93, 0.93, 1)
SUBTEXT = (0.60, 0.60, 0.65, 1)
Window.clearcolor = BG


# ── Helpers ──────────────────────────────────────────────────────────────────
def make_btn(text, callback, bg=None, height=48):
    bg = bg or ACCENT
    btn = Button(
        text=text, size_hint_y=None, height=dp(height),
        background_normal='', background_color=bg,
        color=TEXT, font_size=dp(14), bold=True
    )
    btn.bind(on_press=callback)
    return btn

def card(padding=12):
    box = BoxLayout(orientation='vertical', padding=dp(padding), spacing=dp(6),
                    size_hint_y=None)
    box.bind(minimum_height=box.setter('height'))
    with box.canvas.before:
        Color(*PANEL)
        box._rect = RoundedRectangle(pos=box.pos, size=box.size, radius=[dp(10)])
    box.bind(pos=lambda w, v: setattr(w._rect, 'pos', v),
             size=lambda w, v: setattr(w._rect, 'size', v))
    return box


# ── State ────────────────────────────────────────────────────────────────────
class AppState:
    excel_path = ''
    workbook   = None
    worksheet  = None
    total_rows = 0
    range_start = 1
    range_end   = 100
    current_row = 1

state = AppState()

def load_excel(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    state.workbook  = wb
    state.worksheet = ws
    state.excel_path = path
    state.total_rows = ws.max_row - 1  # subtract header
    return state.total_rows

def get_row_data(row_num):
    ws = state.worksheet
    r  = row_num + 1  # +1 for header
    return {
        'barcode':     ws.cell(r, 1).value or '',
        'codigo':      ws.cell(r, 2).value or '',
        'descripcion': ws.cell(r, 3).value or '',
    }

def save_barcode(row_num, barcode):
    ws = state.worksheet
    ws.cell(row_num + 1, 1, barcode)
    state.workbook.save(state.excel_path)

def export_excel(dest_path):
    state.workbook.save(dest_path)


# ── Screens ───────────────────────────────────────────────────────────────────

class HomeScreen(Screen):
    def __init__(self, **kw):
        super().__init__(**kw)
        self.name = 'home'
        root = BoxLayout(orientation='vertical', padding=dp(16), spacing=dp(12))

        # Title
        root.add_widget(Label(text='Comercial Barrientos',
                              font_size=dp(22), bold=True, color=ACCENT,
                              size_hint_y=None, height=dp(50)))

        # Import card
        ic = card()
        ic.add_widget(Label(text='Archivo Excel', font_size=dp(13),
                            color=SUBTEXT, size_hint_y=None, height=dp(24)))
        self.path_label = Label(text='Sin archivo cargado', font_size=dp(12),
                                color=TEXT, size_hint_y=None, height=dp(24),
                                text_size=(Window.width - dp(60), None),
                                halign='left', valign='middle')
        ic.add_widget(self.path_label)
        ic.add_widget(make_btn('📂  Importar Excel', self.open_import))
        root.add_widget(ic)

        # Range card
        rc = card()
        rc.add_widget(Label(text='Rango de productos', font_size=dp(13),
                            color=SUBTEXT, size_hint_y=None, height=dp(24)))
        row_box = BoxLayout(size_hint_y=None, height=dp(44), spacing=dp(8))
        self.inp_start = TextInput(text='1', multiline=False, input_filter='int',
                                   font_size=dp(16), halign='center',
                                   background_color=(.06,.09,.18,1), foreground_color=TEXT)
        row_box.add_widget(Label(text='del', size_hint_x=None, width=dp(24), color=SUBTEXT))
        row_box.add_widget(self.inp_start)
        row_box.add_widget(Label(text='al', size_hint_x=None, width=dp(24), color=SUBTEXT))
        self.inp_end = TextInput(text='100', multiline=False, input_filter='int',
                                 font_size=dp(16), halign='center',
                                 background_color=(.06,.09,.18,1), foreground_color=TEXT)
        row_box.add_widget(self.inp_end)
        rc.add_widget(row_box)
        rc.add_widget(make_btn('▶  Comenzar a escanear', self.go_scan))
        root.add_widget(rc)

        # Export
        root.add_widget(make_btn('💾  Exportar / Guardar Excel', self.do_export,
                                  bg=(0.06, 0.31, 0.20, 1)))

        root.add_widget(Label())  # spacer
        self.add_widget(root)

    def open_import(self, *_):
        self.manager.current = 'import'

    def go_scan(self, *_):
        if not state.workbook:
            self._toast('Primero importá un archivo Excel'); return
        try:
            s = int(self.inp_start.text)
            e = int(self.inp_end.text)
        except ValueError:
            self._toast('Rango inválido'); return
        if s < 1 or e < s or e > state.total_rows:
            self._toast(f'Rango inválido (máx: {state.total_rows})'); return
        state.range_start = s
        state.range_end   = e
        state.current_row = s
        self.manager.get_screen('scan').load_current()
        self.manager.current = 'scan'

    def do_export(self, *_):
        if not state.workbook:
            self._toast('No hay archivo cargado'); return
        if platform == 'android':
            dest = os.path.join(primary_external_storage_path(),
                                'Download', 'barrientos_export.xlsx')
        else:
            dest = os.path.join(os.path.dirname(state.excel_path), 'barrientos_export.xlsx')
        try:
            export_excel(dest)
            self._toast(f'Guardado en:\n{dest}')
        except Exception as ex:
            self._toast(f'Error: {ex}')

    def set_path_label(self, path):
        name = os.path.basename(path)
        self.path_label.text = name

    def _toast(self, msg):
        popup = Popup(title='', content=Label(text=msg, color=TEXT),
                      size_hint=(0.8, None), height=dp(140),
                      background_color=PANEL)
        popup.open()
        Clock.schedule_once(lambda *_: popup.dismiss(), 3)


class ImportScreen(Screen):
    def __init__(self, **kw):
        super().__init__(**kw)
        self.name = 'import'
        root = BoxLayout(orientation='vertical', padding=dp(12), spacing=dp(10))

        root.add_widget(Label(text='Seleccioná el archivo Excel',
                              font_size=dp(16), bold=True, color=ACCENT,
                              size_hint_y=None, height=dp(40)))

        if platform == 'android':
            start_path = primary_external_storage_path()
        else:
            start_path = os.path.expanduser('~')

        self.fc = FileChooserListView(path=start_path, filters=['*.xlsx', '*.xls'],
                                      background_color=PANEL)
        root.add_widget(self.fc)

        btn_box = BoxLayout(size_hint_y=None, height=dp(48), spacing=dp(8))
        btn_box.add_widget(make_btn('✕ Cancelar', lambda *_: setattr(self.manager, 'current', 'home'),
                                    bg=(0.25, 0.25, 0.28, 1)))
        btn_box.add_widget(make_btn('✓ Cargar', self.load_file))
        root.add_widget(btn_box)
        self.add_widget(root)

    def on_pre_enter(self, *_):
        if platform == 'android':
            request_permissions([Permission.READ_EXTERNAL_STORAGE,
                                  Permission.WRITE_EXTERNAL_STORAGE])

    def load_file(self, *_):
        sel = self.fc.selection
        if not sel:
            return
        path = sel[0]
        try:
            n = load_excel(path)
            home = self.manager.get_screen('home')
            home.set_path_label(path)
            home._toast(f'Cargado: {n} productos')
            self.manager.current = 'home'
        except Exception as ex:
            self.manager.get_screen('home')._toast(f'Error: {ex}')
            self.manager.current = 'home'


class ScanScreen(Screen):
    def __init__(self, **kw):
        super().__init__(**kw)
        self.name = 'scan'
        self._camera = None
        self._scan_event = None
        self._last_code = ''

        root = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(8))

        # Header
        hdr = BoxLayout(size_hint_y=None, height=dp(40))
        hdr.add_widget(make_btn('◀ Volver', self.go_home, bg=(0.2,0.2,0.25,1), height=36))
        self.lbl_progress = Label(text='', color=SUBTEXT, font_size=dp(12))
        hdr.add_widget(self.lbl_progress)
        root.add_widget(hdr)

        # Product card
        pc = card(10)
        self.lbl_row   = Label(text='', color=SUBTEXT, font_size=dp(11),
                               size_hint_y=None, height=dp(18))
        self.lbl_desc  = Label(text='', color=TEXT, font_size=dp(16), bold=True,
                               size_hint_y=None, height=dp(46),
                               text_size=(Window.width - dp(48), None),
                               halign='left', valign='middle')
        self.lbl_cod   = Label(text='', color=SUBTEXT, font_size=dp(11),
                               size_hint_y=None, height=dp(18))
        self.lbl_saved = Label(text='', color=(0.5, 0.8, 0.5, 1), font_size=dp(11),
                               size_hint_y=None, height=dp(18))
        pc.add_widget(self.lbl_row)
        pc.add_widget(self.lbl_desc)
        pc.add_widget(self.lbl_cod)
        pc.add_widget(self.lbl_saved)
        root.add_widget(pc)

        # Camera
        self._cam_box = BoxLayout(size_hint_y=None, height=dp(220))
        root.add_widget(self._cam_box)

        # Manual input
        mi = BoxLayout(size_hint_y=None, height=dp(46), spacing=dp(8))
        self.inp_manual = TextInput(hint_text='Código manual...', multiline=False,
                                    input_filter='int', font_size=dp(14),
                                    background_color=(.06,.09,.18,1), foreground_color=TEXT)
        self.inp_manual.bind(on_text_validate=self.submit_manual)
        mi.add_widget(self.inp_manual)
        mi.add_widget(make_btn('OK', self.submit_manual, height=46))
        root.add_widget(mi)

        # Nav buttons
        nav = BoxLayout(size_hint_y=None, height=dp(46), spacing=dp(8))
        nav.add_widget(make_btn('◀ Ant', self.go_prev, bg=(0.2,0.2,0.25,1)))
        nav.add_widget(make_btn('Saltar ▶', self.go_next_skip, bg=(0.25,0.25,0.28,1)))
        nav.add_widget(make_btn('✓ Listo ▶', self.go_next_done, bg=(0.06,0.31,0.20,1)))
        root.add_widget(nav)

        # Result flash
        self.lbl_flash = Label(text='', color=(0.2, 0.9, 0.4, 1),
                               font_size=dp(14), bold=True,
                               size_hint_y=None, height=dp(30))
        root.add_widget(self.lbl_flash)

        root.add_widget(Label())  # spacer
        self.add_widget(root)

    def on_pre_enter(self, *_):
        self._start_camera()

    def on_pre_leave(self, *_):
        self._stop_camera()

    def _start_camera(self):
        if self._camera:
            return
        try:
            cam = Camera(play=True, resolution=(640, 480), size_hint=(1, 1))
            self._cam_box.add_widget(cam)
            self._camera = cam
            self._scan_event = Clock.schedule_interval(self._process_frame, 0.5)
        except Exception:
            self._cam_box.add_widget(
                Label(text='Cámara no disponible\nUsá el campo manual',
                      color=ACCENT, halign='center'))

    def _stop_camera(self):
        if self._scan_event:
            self._scan_event.cancel()
            self._scan_event = None
        if self._camera:
            self._camera.play = False
            self._cam_box.remove_widget(self._camera)
            self._camera = None

    def _process_frame(self, *_):
        if not self._camera or not self._camera.texture:
            return
        try:
            from pyzbar.pyzbar import decode
            from PIL import Image as PILImage
            import numpy as np
            tex = self._camera.texture
            buf = tex.pixels
            img = PILImage.frombytes('RGBA', tex.size, buf)
            codes = decode(img)
            if codes:
                code = codes[0].data.decode('utf-8')
                if code != self._last_code:
                    self._last_code = code
                    Clock.schedule_once(lambda *_: self._on_scan(code), 0)
                    Clock.schedule_once(lambda *_: setattr(self, '_last_code', ''), 2.5)
        except Exception:
            pass

    def _on_scan(self, code):
        self._save_and_advance(code)

    def submit_manual(self, *_):
        code = self.inp_manual.text.strip()
        if code:
            self.inp_manual.text = ''
            self._save_and_advance(code)

    def _save_and_advance(self, code):
        try:
            save_barcode(state.current_row, code)
            row_data = get_row_data(state.current_row)
            self._flash(f'✓  {row_data["descripcion"] or code}')
            self.go_next_done()
        except Exception as ex:
            self._flash(f'Error: {ex}')

    def load_current(self):
        r    = state.current_row
        done = r - state.range_start
        total = state.range_end - state.range_start + 1
        self.lbl_progress.text = f'{done}/{total}  •  fila {r}'
        data = get_row_data(r)
        self.lbl_row.text   = f'Producto #{r}'
        self.lbl_desc.text  = data['descripcion'] or '— sin descripción —'
        self.lbl_cod.text   = f'Cód: {data["codigo"]}' if data['codigo'] else ''
        self.lbl_saved.text = (f'✓ Ya tiene: {data["barcode"]}' if data['barcode']
                               else '⬜ Sin código')
        self.lbl_flash.text = ''

    def go_prev(self, *_):
        if state.current_row > state.range_start:
            state.current_row -= 1
            self.load_current()

    def go_next_skip(self, *_):
        self._advance()

    def go_next_done(self, *_):
        self._advance()

    def _advance(self):
        if state.current_row >= state.range_end:
            self._finish()
        else:
            state.current_row += 1
            self.load_current()

    def _finish(self):
        popup = Popup(title='',
                      content=Label(text='✅  Rango completado!', color=TEXT,
                                    font_size=dp(16)),
                      size_hint=(0.8, None), height=dp(120),
                      background_color=GREEN)
        popup.open()
        Clock.schedule_once(lambda *_: popup.dismiss(), 2)
        Clock.schedule_once(lambda *_: self.go_home(), 2.1)

    def go_home(self, *_):
        self.manager.current = 'home'

    def _flash(self, msg):
        self.lbl_flash.text = msg
        Clock.schedule_once(lambda *_: setattr(self.lbl_flash, 'text', ''), 2.5)


# ── App ───────────────────────────────────────────────────────────────────────
class BarrientosApp(App):
    def build(self):
        if platform == 'android':
            request_permissions([
                Permission.CAMERA,
                Permission.READ_EXTERNAL_STORAGE,
                Permission.WRITE_EXTERNAL_STORAGE,
            ])
        sm = ScreenManager(transition=SlideTransition())
        sm.add_widget(HomeScreen())
        sm.add_widget(ImportScreen())
        sm.add_widget(ScanScreen())
        return sm

if __name__ == '__main__':
    BarrientosApp().run()
